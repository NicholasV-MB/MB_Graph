import openpyxl

def extract_rows_from_excel(excel_file):
    """
    Method for extract rows from excel file
    @param: excel_file
    @return: excel_data: list of dict with data files
    """
    wb = openpyxl.load_workbook(excel_file)

    # getting a particular sheet by name out of many sheets
    sheet_names = wb.sheetnames
    worksheet = wb[sheet_names[0]]
    excel_data = list()
    # iterating over the rows and
    # getting value from each cell in row
    row_idx = 0
    for row in worksheet.iter_rows():
        if row_idx > 0 and row[0].value!=None and row[1].value!=None and row[2]!=None:
            row_data = {
                "Title": str(row[0].value),
                "Address": str(row[1].value),
                "Duration": float(row[2].value) if row[2].value!=None else 0.5,
                "Info": str(row[3].value) if row[3].value!=None else ""
            }

            excel_data.append(row_data)
        row_idx += 1
    return excel_data
